# -*- coding: utf-8 -*-
import argparse
import os
import re
import zipfile
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side


def split_service(service: str):
    m = re.match(r"^(.*)\s*\((\d+)\)\s*$", str(service).strip())
    if m:
        return m.group(1).strip(), int(m.group(2))
    return str(service).strip(), 1


def safe_sheet_title(s: str) -> str:
    s = str(s or "").strip()
    s = re.sub(r"[:\\/?*\[\]]", "_", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s[:31] or "Лист"


def unique_sheet_title(wb: Workbook, base: str) -> str:
    base_clean = safe_sheet_title(base)
    if base_clean not in wb.sheetnames:
        return base_clean

    i = 2
    while True:
        suffix = f"_{i}"
        allowed = 31 - len(suffix)
        cand = (base_clean[:allowed] + suffix)[:31]
        if cand not in wb.sheetnames:
            return cand
        i += 1


def apply_styles(ws):
    thin = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max_len + 2


def load_rows_column_b(input_file: str) -> list[str]:
    df = pd.read_excel(input_file, header=None, usecols=[1], engine="openpyxl")
    rows = df.iloc[:, 0].dropna().astype(str).tolist()
    rows = [r.replace("\xa0", " ").strip() for r in rows if " - " in str(r)]
    return rows


def build_workbook(data_dict, title_column: str, include_no_nurse: bool) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    fill1 = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill2 = PatternFill(start_color="CCF2CC", end_color="CCF2CC", fill_type="solid")
    bold = Font(bold=True)

    # service -> persons
    for service, persons in data_dict.items():
        ws_title = unique_sheet_title(wb, str(service))
        ws = wb.create_sheet(title=ws_title)

        ws.cell(row=1, column=1, value=title_column).font = bold
        ws.cell(row=1, column=2, value="Количество").font = bold

        row_idx = 2
        for i, (person, subs) in enumerate(persons.items()):
            if not include_no_nurse and person == "Без медсестры":
                continue

            total = int(sum(subs.values()))
            fill = fill1 if i % 2 == 0 else fill2

            c1 = ws.cell(row=row_idx, column=1, value=person)
            c2 = ws.cell(row=row_idx, column=2, value=total)
            c1.fill = fill
            c2.fill = fill
            c1.font = bold
            row_idx += 1

        apply_styles(ws)

    return wb


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)  # ZIP
    args = ap.parse_args()

    rows = load_rows_column_b(args.input)

    # service -> doctor -> nurse -> count
    data_for_doctors = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    # service -> nurse -> doctor -> count
    data_for_nurses = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

    for row in rows:
        row_fixed = str(row).replace("\xa0", " ").replace(" - ", ";")
        parts = row_fixed.split(";", maxsplit=2)
        if len(parts) != 3:
            continue

        doctor, nurse, service = [p.strip() for p in parts]
        nurse = nurse if nurse else "Без медсестры"
        service_name, qty = split_service(service)

        data_for_doctors[service_name][doctor][nurse] += qty
        data_for_nurses[service_name][nurse][doctor] += qty

    # готовим временные файлы рядом с output
    out_dir = os.path.dirname(os.path.abspath(args.output))
    os.makedirs(out_dir, exist_ok=True)

    doctors_xlsx = os.path.join(out_dir, "Диспа_Врачи.xlsx")
    nurses_xlsx = os.path.join(out_dir, "Диспа_Медсестры.xlsx")

    wb_d = build_workbook(data_for_doctors, "Врач", include_no_nurse=True)
    wb_d.save(doctors_xlsx)

    wb_n = build_workbook(data_for_nurses, "Медсестра", include_no_nurse=False)
    wb_n.save(nurses_xlsx)

    # пишем ZIP
    with zipfile.ZipFile(args.output, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.write(doctors_xlsx, arcname="Диспа_Врачи.xlsx")
        zf.write(nurses_xlsx, arcname="Диспа_Медсестры.xlsx")

    # чистим временные xlsx
    try:
        os.remove(doctors_xlsx)
        os.remove(nurses_xlsx)
    except Exception:
        pass


if __name__ == "__main__":
    main()
