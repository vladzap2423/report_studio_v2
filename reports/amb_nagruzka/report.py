# -*- coding: utf-8 -*-
import argparse
import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill


def apply_borders(ws):
    thin = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for c in row:
            c.border = thin
            c.alignment = Alignment(horizontal="left", vertical="center")


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max_len + 2


def merge_block(ws, start_row, end_row, col):
    if end_row > start_row:
        ws.merge_cells(start_row=start_row, start_column=col,
                       end_row=end_row, end_column=col)
    ws.cell(row=start_row, column=col).alignment = Alignment(
        vertical="center", horizontal="left"
    )


def load_rows_column_b(input_file: str) -> list[str]:
    """
    Исходники у тебя читались как: usecols=[1], header=None, skiprows=3.
    Для устойчивости: читаем колонку B без header, но умеем пропускать пустые.
    """
    df = pd.read_excel(input_file, header=None, usecols=[1], engine="openpyxl")
    rows = df[1].dropna().astype(str).tolist()
    # часто первые строки — шапка/пустое, фильтруем по наличию " - "
    rows = [r.replace("\xa0", " ").strip() for r in rows if " - " in str(r)]
    return rows


def parse(records: list[str]):
    records_doctors = []
    records_nurses = []

    for row in records:
        s = str(row).replace("\xa0", " ").strip()
        parts = s.split(" - ", 2)
        if len(parts) < 3:
            continue

        doctor = parts[0].strip()
        nurses_block = parts[1].strip()
        service_part = parts[2].strip()

        # список медсестёр
        if nurses_block:
            nurse_list = [n.strip() for n in nurses_block.split(";") if n.strip()]
        else:
            nurse_list = ["Без медсестры"]

        nurse_list = [
            re.sub(r"\(\s*-\s*\)", "", n).replace("  ", " ").strip() or "Без медсестры"
            for n in nurse_list
        ]

        m = re.search(r"\((\d+)\)\s*$", service_part)
        count = int(m.group(1)) if m else 0
        service = service_part[:m.start()].strip() if m else service_part

        records_doctors.append({"Doctor": doctor, "Service": service, "Count": count})

        for nurse in nurse_list:
            records_nurses.append({"Nurse": nurse, "Service": service, "Count": count})

    df_doctors = pd.DataFrame(records_doctors)
    df_nurses = pd.DataFrame(records_nurses)

    by_doctor = (
        df_doctors.groupby(["Doctor", "Service"], as_index=False)["Count"]
        .sum()
        .sort_values(["Doctor", "Service"], kind="mergesort")
    )

    by_nurse = (
        df_nurses[df_nurses["Nurse"] != "Без медсестры"]
        .groupby(["Nurse", "Service"], as_index=False)["Count"]
        .sum()
        .sort_values(["Nurse", "Service"], kind="mergesort")
    )

    return by_doctor, by_nurse


def write_grouped_table(ws, df: pd.DataFrame, group_col: str, title_col: str):
    headers = [title_col, "Услуга", "Количество"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    fills = [
        PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        PatternFill(start_color="CCF2CC", end_color="CCF2CC", fill_type="solid"),
    ]

    row_idx = 2
    groups = list(df.groupby(group_col, sort=False))

    for i, (name, group) in enumerate(groups):
        total = int(group["Count"].sum())
        first_row = row_idx
        fill = fills[i % 2]

        for _, row in group.iterrows():
            ws.cell(row=row_idx, column=2, value=row["Service"]).fill = fill
            ws.cell(row=row_idx, column=3, value=int(row["Count"])).fill = fill
            row_idx += 1

        ws.cell(row=first_row, column=1, value=f"{name} ({total})").fill = fill
        merge_block(ws, first_row, row_idx - 1, 1)

    apply_borders(ws)
    auto_width(ws)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Входной Excel (xlsx/xls)")
    ap.add_argument("--output", required=True, help="Выходной Excel (xlsx)")
    args = ap.parse_args()

    rows = load_rows_column_b(args.input)
    by_doctor, by_nurse = parse(rows)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "По врачам"
    write_grouped_table(ws1, by_doctor, "Doctor", "Врач")

    ws2 = wb.create_sheet(title="По медсестрам")
    write_grouped_table(ws2, by_nurse, "Nurse", "Медсестра")

    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    wb.save(args.output)


if __name__ == "__main__":
    main()
