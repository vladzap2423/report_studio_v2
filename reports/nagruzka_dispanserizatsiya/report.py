import re
import pandas as pd
from collections import defaultdict
from tkinter import Tk, filedialog

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter


def split_service(service: str):
    match = re.match(r"^(.*)\s*\((\d+)\)\s*$", service.strip())
    if match:
        name = match.group(1).strip()
        count = int(match.group(2))
        return name, count
    return service.strip(), 1


def normalize_text(value: str) -> str:
    return str(value).replace("\xa0", " ").strip()


def style_sheet(ws, headers_count: int, data_rows_count: int):
    thin = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    header_fill = PatternFill(start_color="F3E5AB", end_color="F3E5AB", fill_type="solid")
    bold_font = Font(bold=True)
    value_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for col in range(1, headers_count + 1):
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin

    for col in range(1, headers_count + 1):
        c = ws.cell(row=2, column=col)
        c.fill = header_fill
        c.border = thin
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=3, max_row=2 + data_rows_count, min_col=1, max_col=headers_count):
        for cell in row:
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.column != 1 and isinstance(cell.value, (int, float)) and cell.value > 0:
                cell.fill = value_fill

    for col_idx in range(1, headers_count + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 18), 40)

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 10
    ws.freeze_panes = "B3"


def fill_sheet(ws, title_name: str, row_dict: dict, service_columns: list):
    ws["A1"] = title_name

    # Заголовки услуг
    for i, service in enumerate(service_columns, start=2):
        ws.cell(row=1, column=i, value=service)

    # Новый правый столбец "Всего"
    total_col = len(service_columns) + 2
    ws.cell(row=1, column=total_col, value="Всего")

    people = sorted(row_dict.keys(), key=lambda x: x.lower())

    # Данные по людям
    for row_idx, person in enumerate(people, start=3):
        ws.cell(row=row_idx, column=1, value=person)

        row_total = 0
        for col_idx, service in enumerate(service_columns, start=2):
            value = row_dict[person].get(service, 0)
            ws.cell(row=row_idx, column=col_idx, value=value)
            row_total += value

        # Сумма по строке справа
        ws.cell(row=row_idx, column=total_col, value=row_total)

    # Нижняя строка "Всего"
    total_row = len(people) + 3
    ws.cell(row=total_row, column=1, value="Всего")

    # Суммы по каждому столбцу услуг
    grand_total = 0
    for col_idx, service in enumerate(service_columns, start=2):
        col_total = 0
        for person in people:
            col_total += row_dict[person].get(service, 0)

        ws.cell(row=total_row, column=col_idx, value=col_total)
        grand_total += col_total

    # Правый нижний угол — общий итог
    ws.cell(row=total_row, column=total_col, value=grand_total)

    style_sheet(ws, headers_count=total_col, data_rows_count=len(people) + 1)

Tk().withdraw()
input_file = filedialog.askopenfilename(
    title="Выберите Excel-файл",
    filetypes=[("Excel files", "*.xlsx *.xls")]
)

if not input_file:
    print("Файл не выбран. Выход.")
    raise SystemExit

df = pd.read_excel(input_file, header=None, usecols=[1], skiprows=3)
rows = df[1].dropna().tolist()

doctors = defaultdict(lambda: defaultdict(int))
nurses = defaultdict(lambda: defaultdict(int))


found_services = set()

for row in rows:
    row_fixed = normalize_text(row).replace(" - ", ";")
    parts = row_fixed.split(";", maxsplit=2)

    if len(parts) != 3:
        continue

    doctor, nurse, service = [normalize_text(p) for p in parts]

    if not doctor:
        continue

    if not nurse:
        nurse = "Без медсестры"

    service_name, count = split_service(service)

    doctors[doctor][service_name] += count
    nurses[nurse][service_name] += count
    found_services.add(service_name)

service_columns = sorted(found_services, key=str.lower)

if not service_columns:
    print("После исключения столбцов из списка не осталось услуг для вывода.")
    raise SystemExit

wb = Workbook()

ws_doctors = wb.active
ws_doctors.title = "Врачи"
fill_sheet(ws_doctors, "ФИО Врача", doctors, service_columns)

ws_nurses = wb.create_sheet("Медсестры")
fill_sheet(ws_nurses, "ФИО Медсестры", nurses, service_columns)

out_file = "Диспа_сводный_отчет.xlsx"
wb.save(out_file)

print(f"Сохранён файл: {out_file}")