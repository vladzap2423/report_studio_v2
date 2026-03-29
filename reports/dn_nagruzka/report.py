# -*- coding: utf-8 -*-
import argparse
import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side


def load_rows_column_b(input_file: str) -> list[str]:
    df = pd.read_excel(input_file, header=None, usecols=[1], engine="openpyxl")
    rows = df[1].dropna().astype(str).tolist()
    rows = [r.replace("\xa0", " ").strip() for r in rows if " - " in str(r)]
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Входной Excel (xlsx/xls)")
    ap.add_argument("--output", required=True, help="Выходной Excel (xlsx)")
    args = ap.parse_args()

    rows = load_rows_column_b(args.input)

    doctor_totals = {}
    doctor_cases = {}

    for row in rows:
        parts = row.split(" - ", maxsplit=1)
        if len(parts) != 2:
            continue

        doctor = parts[0].strip()
        rest = parts[1].strip()

        m = re.match(r"(\d+)\s*\((\d+)\)", rest)
        if not m:
            continue

        kd = int(m.group(1))
        kol_vo_sluch = int(m.group(2))

        if kd > 10:
            kd = 10
        elif 5 < kd <= 10:
            kd = kd - 2

        kol_vo_kd = kd * kol_vo_sluch

        doctor_totals[doctor] = doctor_totals.get(doctor, 0) + kol_vo_kd
        doctor_cases[doctor] = doctor_cases.get(doctor, 0) + kol_vo_sluch

    wb = Workbook()
    ws = wb.active
    ws.title = "Итог"

    ws.append(["Врач", "Сумма койка-дней"])

    fill1 = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill2 = PatternFill(start_color="CCF2CC", end_color="CCF2CC", fill_type="solid")

    # стабильная сортировка по ФИО
    doctors = sorted(doctor_totals.keys(), key=lambda x: (str(x).lower(), str(x)))

    for i, doctor in enumerate(doctors, start=2):
        total_kd = int(doctor_totals[doctor])
        total_cases = int(doctor_cases.get(doctor, 0))
        ws.append([f"{doctor} ({total_cases})", total_kd])

        fill = fill1 if i % 2 == 0 else fill2
        ws[f"A{i}"].fill = fill
        ws[f"B{i}"].fill = fill

    alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = alignment
            cell.border = border

    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2

    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    wb.save(args.output)


if __name__ == "__main__":
    main()
