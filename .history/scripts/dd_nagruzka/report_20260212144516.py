# -*- coding: utf-8 -*-
import argparse
import os
import re
import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side


def split_service(service: str):
    m = re.match(r"^(.*)\s*\((\d+)\)\s*$", service.strip())
    if m:
        return m.group(1).strip(), int(m.group(2))
    return service.strip(), 1


def apply_styles(ws):
    thin = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max_len + 2


def load_rows_column_b(input_file: str) -> list[str]:
    df = pd.read_excel(input_file, header=None, usecols=[1], engine="openpyxl")
    rows = df[1].dropna().astype(str).tolist()
    rows = [r.replace("\xa0", " ").strip() for r in rows if " - " in str(r)]
    return rows


def create_report(data_dict, out_file, title_column, include_no_nurse=True):
    wb = Workbook()
    wb.remove(wb.active)

    fill1 = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill2 = PatternFill(start_color="CCF2CC", end_color="CCF2CC", fill_type="solid")
    bold_font = Font(bold=True)

    for service, persons in data_dict.items():
        ws = wb.create_sheet(title=str(service)[:31])
        ws.cell(row=1, column=1, value=title_column).font = bold_font
        ws.cell(row=1, column=2, value="Количество").font = bold_font

        row_idx = 2
        for i, (person, subs) in enumerate(persons.items()):
            if not include_no_nurse and person == "Без медсестры":
                continue

            fill = fill1 if i % 2 == 0 else fill2
            total = int(sum(subs.values()))

            ws.cell(row=row_idx, column=1, value=person).fill = fill
            ws.cell(row=row_idx, column=2, value=total).fill = fill
            ws.cell(row=row_idx, column=1).font = bold_font
            row_idx += 1

        apply_styles(ws)

    os.makedirs(os.path.dirname(os.path.abspath(out_file)), exist_ok=True)
    wb.save(out_file)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Входной Excel (xlsx/xls)")
    ap.add_argument("--output", required=True, help="Выходной Excel (xlsx)")
    args = ap.parse_args()

    rows = load_rows_column_b(args.input)

    data_for_doctors = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    data_for_nurses = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

    for row in rows:
        row_fixed = str(row).replace("\xa0", " ").replace(" - ", ";")
        parts = row_fixed.split(";", maxsplit=2)
        if len(parts) != 3:
            continue
        doctor, nurse, service = [p.strip() for p in parts]
        nurse = nurse if nurse else "Без медсестры"
        service_name, kol_vo = split_service(service)

        data_for_doctors[service_name][doctor][nurse] += kol_vo
        data_for_nurses[service_name][nurse][doctor] += kol_vo

    # один output файл: создадим книгу, а внутри 2 “набора листов” через префиксы
    # чтобы не было конфликтов имён листов
    tmp_doctors = args.output + ".doctors.tmp.xlsx"
    tmp_nurses = args.output + ".nurses.tmp.xlsx"

    create_report(data_for_doctors, tmp_doctors, "Врач", include_no_nurse=True)
    create_report(data_for_nurses, tmp_nurses, "Медсестра", include_no_nurse=False)

    # склеим в один файл (два набора листов)
    from openpyxl import load_workbook
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    wb_d = load_workbook(tmp_doctors)
    for ws in wb_d.worksheets:
        new = wb_out.create_sheet(title=("Врачи_" + ws.title)[:31])
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                new[cell.coordinate].value = cell.value
                new[cell.coordinate]._style = cell._style
        new.column_dimensions.update(ws.column_dimensions)
        new.row_dimensions.update(ws.row_dimensions)

    wb_n = load_workbook(tmp_nurses)
    for ws in wb_n.worksheets:
        new = wb_out.create_sheet(title=("Мс_" + ws.title)[:31])
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                new[cell.coordinate].value = cell.value
                new[cell.coordinate]._style = cell._style
        new.column_dimensions.update(ws.column_dimensions)
        new.row_dimensions.update(ws.row_dimensions)

    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    wb_out.save(args.output)

    # чистим tmp
    try:
        os.remove(tmp_doctors)
        os.remove(tmp_nurses)
    except Exception:
        pass


if __name__ == "__main__":
    main()
