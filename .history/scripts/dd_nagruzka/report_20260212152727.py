# -*- coding: utf-8 -*-
import argparse
import os
import re
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side


def split_service(service: str):
    m = re.match(r"^(.*)\s*\((\d+)\)\s*$", service.strip())
    if m:
        return m.group(1).strip(), int(m.group(2))
    return service.strip(), 1


def safe_sheet_title(s: str) -> str:
    # Excel: <= 31 символ, нельзя: : \ / ? * [ ]
    s = str(s or "").strip()
    s = re.sub(r"[:\\/?*\[\]]", "_", s)
    s = s[:31].strip()
    return s or "Лист"


def apply_styles(ws):
    thin = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # автоширина
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max_len + 2


def load_rows_column_b(input_file: str) -> list[str]:
    df = pd.read_excel(input_file, header=None, usecols=[1], engine="openpyxl")
    rows = df[1].dropna().astype(str).tolist()
    rows = [r.replace("\xa0", " ").strip() for r in rows if " - " in str(r)]
    return rows


def write_service_sheet(ws, persons_dict: dict, title_column: str, include_no_nurse: bool):
    fill1 = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill2 = PatternFill(start_color="CCF2CC", end_color="CCF2CC", fill_type="solid")
    bold = Font(bold=True)

    ws.cell(row=1, column=1, value=title_column).font = bold
    ws.cell(row=1, column=2, value="Количество").font = bold

    row_idx = 2
    for i, (person, subs) in enumerate(persons_dict.items()):
        if not include_no_nurse and person == "Без медсестры":
            continue

        fill = fill1 if i % 2 == 0 else fill2
        total = int(sum(subs.values()))

        c1 = ws.cell(row=row_idx, column=1, value=person)
        c2 = ws.cell(row=row_idx, column=2, value=total)
        c1.fill = fill
        c2.fill = fill
        c1.font = bold
        row_idx += 1

    apply_styles(ws)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Входной Excel (xlsx/xls)")
    ap.add_argument("--output", required=True, help="Выходной Excel (xlsx)")
    args = ap.parse_args()

    rows = load_rows_column_b(args.input)

    # service -> doctor -> nurse -> count
    data_for_doctors = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    # service -> nurse -> doctor -> count
    data_for_nurses = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

    for row in rows:
        row_fixed = str(row).replace("\xa0", " ").replace(" - ", ";")
        parts = row_fixed.split(";", maxsplit=2)
        if len(parts) != 3:
            continue

        doctor, nurse, service = [p.strip() for p in parts]
        nurse = nurse if nurse else "Без медсестры"

        service_name, qty = split_service(service)

        data_for_doctors[service_name][doctor][nurse] += qty
        data_for_nurses[service_name][nurse][doctor] += qty

    # Пишем сразу в одну книгу, без переносов стилей между wb
    wb = Workbook()
    wb.remove(wb.active)

    # Листы по врачам
    for service, persons in data_for_doctors.items():
        title = safe_sheet_title("Врачи_" + str(service))
        ws = wb.create_sheet(title=title)
        write_service_sheet(ws, persons, "Врач", include_no_nurse=True)

    # Листы по медсестрам (без "Без медсестры")
    for service, persons in data_for_nurses.items():
        title = safe_sheet_title("Мс_" + str(service))
        ws = wb.create_sheet(title=title)
        write_service_sheet(ws, persons, "Медсестра", include_no_nurse=False)

    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    wb.save(args.output)


if __name__ == "__main__":
    main()
