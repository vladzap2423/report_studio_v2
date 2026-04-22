from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook, Workbook


OUTPUT_SHEET = "Результат"

FIELD_MAP = {
    "ФИО пациента": "ФИО пациента",
    "Врач": "Врач",
    "Участок": "Участок",
    "Данные о льготах": "Тип льготы",
    "Категория льготы": "Категория льготы",
    "Диагноз": "Диагноз",
    "Наличие инвалидности(группа)": "Инвалидность",
}

MED_COLUMNS = [
    "ЖНВЛП",
    "МНН",
    "Торговое наименование",
    "Дозировка",
    "Форма",
    "Потребность в месяц",
    "В упаковке",
    "Упаковок",
    "Всего штук",
]


# ---------------- utils ----------------

def normalize(value):
    return str(value).strip() if value else ""


def safe_get(values, idx):
    return values[idx] if idx < len(values) else ""


def clean_field_name(name: str) -> str:
    return re.sub(r"^\d+\.\s*", "", name).strip()


def parse_field(text: str):
    parts = text.split(":", 1)
    if len(parts) != 2:
        return "", ""

    raw_key = clean_field_name(parts[0])
    value = parts[1].replace("\n", " ").strip()

    key = FIELD_MAP.get(raw_key, raw_key)
    return key, value


# ---------------- detection ----------------

def is_new_patient(row):
    return bool(re.match(r"^\s*\d+\s*$", row[0]))


def is_patient_field(text: str):
    if not text:
        return False
    return bool(re.match(r"^\s*\d+\.\s*", text))


def is_med_header(row):
    joined = " ".join(row).lower()
    return "мнн" in joined and "жнвлп" in joined


def is_med_row(row):
    return len(row) > 2 and bool(re.match(r"^\d+$", row[2]))


def is_empty_med(med: dict) -> bool:
    return not (med.get("МНН") or med.get("Дозировка"))


# ---------------- parser ----------------

def parse_input_sheet(ws):
    patients = []

    current_patient: Dict[str, str] = {}
    current_meds: List[Dict[str, str]] = []
    mode = "header"

    for row in ws.iter_rows(values_only=True):
        values = [normalize(v) for v in row]

        if not any(values):
            continue

        # новый пациент
        if is_new_patient(values):
            if current_patient or current_meds:
                patients.append((current_patient, current_meds))

            current_patient = {}
            current_meds = []
            mode = "header"

        # --------------------
        # ПОЛЯ ПАЦИЕНТА (FIX)
        # --------------------
        for cell in values:
            if is_patient_field(cell):
                key, val = parse_field(cell)
                if key:
                    current_patient[key] = val

        # заголовок лекарств
        if is_med_header(values):
            mode = "meds"
            continue

        # лекарства
        if mode == "meds" and is_med_row(values):
            med = {
                "ЖНВЛП": safe_get(values, 3),
                "МНН": safe_get(values, 4),
                "Торговое наименование": safe_get(values, 5),
                "Дозировка": safe_get(values, 6),
                "Форма": safe_get(values, 7),
                "Потребность в месяц": safe_get(values, 8),
                "В упаковке": safe_get(values, 9),
                "Упаковок": safe_get(values, 10),
                "Всего штук": safe_get(values, 11),
            }

            if not is_empty_med(med):
                current_meds.append(med)

    if current_patient or current_meds:
        patients.append((current_patient, current_meds))

    return patients


# ---------------- flatten ----------------

def flatten(patients):
    rows = []

    for patient, meds in patients:
        fio = patient.get("ФИО пациента", "")

        if not meds:
            rows.append(patient)
            continue

        for med in meds:
            row = {"ФИО пациента": fio}
            row.update(patient)
            row.update(med)
            rows.append(row)

    return rows


# ---------------- write ----------------

def write_output(path: Path, rows: List[Dict[str, str]]):
    wb = Workbook()
    ws = wb.active
    ws.title = OUTPUT_SHEET

    if not rows:
        wb.save(path)
        return

    keys = set().union(*(r.keys() for r in rows))

    patient_keys = sorted([k for k in keys if k not in MED_COLUMNS and k != "ФИО пациента"])
    header = ["ФИО пациента"] + patient_keys + MED_COLUMNS

    for col, key in enumerate(header, 1):
        ws.cell(1, col, key)

    for r_idx, row in enumerate(rows, 2):
        for c_idx, key in enumerate(header, 1):
            ws.cell(r_idx, c_idx, row.get(key, ""))

    wb.save(path)


# ---------------- main ----------------

def build_report(input_path: str | Path, output_path: str | Path) -> int:
    wb = load_workbook(input_path, data_only=True)
    ws = wb.active

    patients = parse_input_sheet(ws)
    rows = flatten(patients)

    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    write_output(output_file, rows)

    return len(rows)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)

    args = parser.parse_args()

    row_count = build_report(args.input, args.output)
    print(f"Отчёт успешно сохранён: {args.output}")
    print(f"Готово. Строк: {row_count}")


if __name__ == "__main__":
    main()
