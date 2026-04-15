import re
import pandas as pd
from tkinter import Tk, filedialog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

# =========================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# =========================
def apply_borders(ws):
    thin = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for c in row:
            c.border = thin
            c.alignment = Alignment(horizontal="left", vertical="center")


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max_len + 2


def merge_block(ws, start_row, end_row, col):
    """Объединяет ячейки в колонке col с выравниванием по центру"""
    if end_row > start_row:
        ws.merge_cells(start_row=start_row, start_column=col,
                       end_row=end_row, end_column=col)
    ws.cell(row=start_row, column=col).alignment = Alignment(
        vertical="center", horizontal="left"
    )


# =========================
# ВЫБОР ФАЙЛА
# =========================
Tk().withdraw()
input_file = filedialog.askopenfilename(
    title="Выберите Excel-файл (Врач - Медсестра - Услуга (число))",
    filetypes=[("Excel files", "*.xlsx *.xls")],
)
if not input_file:
    print("Файл не выбран. Выход.")
    raise SystemExit


# =========================
# ЧТЕНИЕ И ПАРСИНГ
# =========================
df = pd.read_excel(input_file, header=None, usecols=[1], skiprows=3)
rows = df[1].dropna().tolist()

records_doctors = []
records_nurses = []

for row in rows:
    s = str(row).replace("\xa0", " ").strip()
    parts = s.split(" - ", 2)
    if len(parts) < 3:
        continue

    doctor = parts[0].strip()
    nurses_block = parts[1].strip()
    service_part = parts[2].strip()

    # список медсестёр
    if nurses_block:
        nurse_list = [n.strip() for n in nurses_block.split(";") if n.strip()]
    else:
        nurse_list = ["Без медсестры"]

    # чистим "(-)" и лишние пробелы
    nurse_list = [
        re.sub(r"\(\s*-\s*\)", "", n).replace("  ", " ").strip() or "Без медсестры"
        for n in nurse_list
    ]

    # количество — из последних скобок
    m = re.search(r"\((\d+)\)\s*$", service_part)
    count = int(m.group(1)) if m else 0
    service = service_part[:m.start()].strip() if m else service_part

    # врач — только одна запись
    records_doctors.append({
        "Doctor": doctor,
        "Service": service,
        "Count": count
    })

    # медсёстры — каждая получает полное количество
    for nurse in nurse_list:
        records_nurses.append({
            "Nurse": nurse,
            "Service": service,
            "Count": count
        })


df_doctors = pd.DataFrame(records_doctors)
df_nurses = pd.DataFrame(records_nurses)


# =========================
# АГРЕГАЦИЯ
# =========================
by_doctor = (
    df_doctors.groupby(["Doctor", "Service"], as_index=False)["Count"]
    .sum()
    .sort_values(["Doctor", "Service"])
)

by_nurse = (
    df_nurses[df_nurses["Nurse"] != "Без медсестры"]
    .groupby(["Nurse", "Service"], as_index=False)["Count"]
    .sum()
    .sort_values(["Nurse", "Service"])
)


# =========================
# СОЗДАНИЕ ТАБЛИЦ
# =========================
def create_report(df, group_col, title_col, out_file, sheet_title):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    headers = [title_col, "Услуга", "Количество"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    fills = [
        PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        PatternFill(start_color="CCF2CC", end_color="CCF2CC", fill_type="solid"),
    ]

    row_idx = 2
    groups = list(df.groupby(group_col))

    for i, (name, group) in enumerate(groups):
        total = group["Count"].sum()
        first_row = row_idx
        fill = fills[i % 2]

        for _, row in group.iterrows():
            ws.cell(row=row_idx, column=2, value=row["Service"]).fill = fill
            ws.cell(row=row_idx, column=3, value=row["Count"]).fill = fill
            row_idx += 1

        ws.cell(row=first_row, column=1, value=f"{name} ({total})").fill = fill
        merge_block(ws, first_row, row_idx - 1, 1)

    apply_borders(ws)
    auto_width(ws)
    wb.save(out_file)
    print(f"Файл сохранён: {out_file}")


# =========================
# СОХРАНЕНИЕ ОТЧЁТОВ
# =========================
create_report(by_doctor, "Doctor", "Врач",
              "Посещения_по_врачам.xlsx", "По врачам")
create_report(by_nurse, "Nurse", "Медсестра",
              "Посещения_по_медсестрам.xlsx", "По медсестрам")
