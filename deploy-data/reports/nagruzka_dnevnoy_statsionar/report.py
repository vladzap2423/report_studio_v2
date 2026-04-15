import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from tkinter import Tk,  filedialog


Tk().withdraw()
input_file = filedialog.askopenfilename(
    title="Выберите Excel-файл",
    filetypes=[("Excel files", "*.xlsx *.xls")]
)

if not input_file:
    print("Файл не выбран. Выход.")
    exit()

output_file = "Дневной.xlsx"

df = pd.read_excel(input_file, header=None, usecols=[1], skiprows=3)
rows = df[1].dropna().tolist()


doctor_totals = {}
doctor_cases = {}

for row in rows:
    parts = row.split(" - ", maxsplit=1)
    if len(parts) == 2:
        doctor = parts[0].strip()
        rest = parts[1].strip()

        match = re.match(r"(\d+)\s*\((\d+)\)", rest)
        if match:
            kd = int(match.group(1))
            kol_vo_sluch = int(match.group(2))
            if kd > 10:
                kd = 10
            elif 5 < kd <= 10:
                kd = kd - 2
            else:
                pass

            kol_vo_kd = kd * kol_vo_sluch

            # Суммируем kol_vo_kd
            doctor_totals[doctor] = doctor_totals.get(doctor, 0) + kol_vo_kd
            # Суммируем kol_vo_sluch
            doctor_cases[doctor] = doctor_cases.get(doctor, 0) + kol_vo_sluch


wb = Workbook()
ws = wb.active
ws.title = "Итог"

ws.append(["Врач", "Сумма койка-дней"])

fill1 = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
fill2 = PatternFill(start_color="CCF2CC", end_color="CCF2CC", fill_type="solid")


for i, doctor in enumerate(doctor_totals.keys(), start=2):
    total_kd = doctor_totals[doctor]
    total_cases = doctor_cases[doctor]
    ws.append([f"{doctor} ({total_cases})", total_kd])
    fill = fill1 if i % 2 == 0 else fill2
    ws[f"A{i}"].fill = fill
    ws[f"B{i}"].fill = fill

alignment = Alignment(horizontal="center", vertical="center")
border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
    for cell in row:
        cell.alignment = alignment
        cell.border = border

for col in ws.columns:
    max_length = 0
    column_letter = col[0].column_letter
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[column_letter].width = max_length + 2

wb.save(output_file)
print(f"Данные успешно записаны в {output_file}")
