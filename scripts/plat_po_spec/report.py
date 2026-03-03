# -*- coding: utf-8 -*-
"""
Отчёт с группировкой по профилю ИЗ БАЗЫ data.db.
Медикаменты и Профиль берутся из таблицы services по коду ОК МУ.
"""
import argparse
import os
import sqlite3

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Путь к базе — относительно папки scripts/dogovor
DB_PATH = os.path.normpath(os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),  # поднимаемся из dogovor → scripts
    "data.db"
))

SHEET_MEDCOM = "Медкомиссии"
SHEET_OTHER = "Остальное"

# Входные колонки из Excel
VID_COL = "Вид поступления"
VID_PROF = "профосмотр"  # нижний регистр для сравнения

COL_CODE = "Код ОКМУ"
COL_SERVICE = "Услуга"
COL_STATE = "Состояние"
COL_DATE = "Дата"
COL_SUM = "Сумма"
COL_FIO = "ФИО"
COL_SPEC = "Специалист/Ресурс.Выполнение"
COL_EXTRA = "Понятие dopPers"
COL_TARPLAN = "Тар.план"
COL_DOGOVOR = "Договор на оплату"

# Выходные колонки (порядок)
OUT_HEADERS = [
    "ФИО",
    "Услуга",
    "Состояние",
    "Дата",
    "Количество услуг",
    "Сумма по тарифу",
    "Медикаменты",
    "Сумма для распределения",
    "Специалист/Ресурс.Выполнение",
    "Персонал. Дополнительный персонал/ресурсы",
]

COL_WIDTHS = [34, 65, 20, 16, 16, 16, 22, 28, 34, 34]


def get_db_connection():
    if not os.path.exists(DB_PATH):
        raise FileNotFoundError(f"База данных не найдена: {DB_PATH}\nПроверьте путь: {DB_PATH}")
    return sqlite3.connect(DB_PATH)


def load_db_maps() -> tuple[dict[str, float], dict[str, str]]:
    """Загружает медикаменты и профили из data.db"""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT "Код_ОК_МУ", "Медикаменты", "Профиль"
        FROM services
        WHERE "Код_ОК_МУ" IS NOT NULL AND "Код_ОК_МУ" != ''
    """)
    rows = cursor.fetchall()
    conn.close()

    meds_map = {}
    profile_map = {}

    for code, meds, prof in rows:
        code_str = str(code).strip()
        if code_str:
            meds_map[code_str] = float(meds) if meds is not None else 0.0
            profile_map[code_str] = str(prof).strip() if prof else "Без профиля"

    return meds_map, profile_map


def clean_text_series(s: pd.Series) -> pd.Series:
    s = s.astype("string").str.strip()
    s = s.replace({"nan": pd.NA, "None": pd.NA, "": pd.NA})
    return s


def normalize_str(x) -> str:
    if pd.isna(x) or x is None:
        return ""
    return str(x).strip().lower()


def load_input(path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = None
    for r in range(1, min(50, ws.max_row) + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(v == COL_CODE for v in vals if v):
            header_row = r
            break

    if header_row is None:
        raise RuntimeError(f"Не найдена строка заголовков с колонкой '{COL_CODE}'")

    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    data = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {headers[c-1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        if all(v is None for v in row.values()):
            continue
        data.append(row)

    wb.close()

    df = pd.DataFrame(data)

    # Обязательные колонки
    required = [COL_CODE, COL_SERVICE, COL_SUM]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"В файле отсутствуют обязательные столбцы: {', '.join(missing)}")

    df[COL_SUM] = pd.to_numeric(df[COL_SUM], errors="coerce").fillna(0)

    for col in [VID_COL, COL_FIO, COL_SERVICE, COL_CODE, COL_STATE,
                COL_SPEC, COL_EXTRA, COL_TARPLAN, COL_DOGOVOR]:
        if col in df.columns:
            df[col] = clean_text_series(df[col])

    if COL_DATE in df.columns:
        df[COL_DATE] = pd.to_datetime(df[COL_DATE], errors="coerce")

    return df[df[COL_CODE].notna()].copy()


def split_by_vid(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    if VID_COL not in df.columns:
        return df.iloc[0:0].copy(), df.copy()
    vid_norm = df[VID_COL].apply(normalize_str)
    return df[vid_norm == VID_PROF].copy(), df[vid_norm != VID_PROF].copy()


def aggregate_for_patient_style(df: pd.DataFrame, meds_map: dict, profile_map: dict) -> pd.DataFrame:
    df = df.copy()

    # Профиль берём из базы
    df["Профиль"] = df[COL_CODE].astype(str).str.strip().map(profile_map).fillna("Без профиля")

    g = df.groupby(
        ["Профиль", COL_FIO, COL_CODE, COL_SERVICE, COL_STATE, COL_DATE, COL_SPEC, COL_EXTRA],
        as_index=False,
        dropna=False
    ).agg(
        **{
            "Количество услуг": (COL_CODE, "size"),
            "Сумма по тарифу": (COL_SUM, "sum"),
        }
    )

    # Медикаменты из базы
    per_service = g[COL_CODE].astype(str).str.strip().map(meds_map).fillna(0.0)
    g["Медикаменты"] = (per_service * g["Количество услуг"]).round(2)
    g["Сумма для распределения"] = (g["Сумма по тарифу"] - g["Медикаменты"]).round(2)

    out = pd.DataFrame({
        "Профиль": g["Профиль"],
        "ФИО": g[COL_FIO],
        "Услуга": g[COL_SERVICE],
        "Состояние": g[COL_STATE],
        "Дата": g[COL_DATE],
        "Количество услуг": g["Количество услуг"].astype(int),
        "Сумма по тарифу": g["Сумма по тарифу"].round(2),
        "Медикаменты": g["Медикаменты"].round(2),
        "Сумма для распределения": g["Сумма для распределения"].round(2),
        "Специалист/Ресурс.Выполнение": g[COL_SPEC],
        "Персонал. Дополнительный персонал/ресурсы": g[COL_EXTRA],
    })

    out.sort_values(["Профиль", "ФИО", "Дата", "Услуга"], inplace=True)
    return out


def write_profile_header(ws, row_idx: int, profile_name: str, ncols: int) -> int:
    title = (profile_name or "Без профиля").strip()
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ncols)
    cell = ws.cell(row_idx, 1, value=title)
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row_idx].height = 28
    return row_idx + 1


def write_patient_like_sheet(ws, df_part: pd.DataFrame, meds_map: dict, profile_map: dict):
    table_df = aggregate_for_patient_style(df_part, meds_map, profile_map)
    ncols = len(OUT_HEADERS)

    # Заголовки
    for c, h in enumerate(OUT_HEADERS, 1):
        cell = ws.cell(1, c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    r = 2
    for profile, df_prof in table_df.groupby("Профиль", sort=False):
        r = write_profile_header(ws, r, profile, ncols)

        for _, row in df_prof.iterrows():
            values = [
                row["ФИО"],
                row["Услуга"],
                row["Состояние"],
                row["Дата"],
                int(row["Количество услуг"]),
                float(row["Сумма по тарифу"]),
                float(row["Медикаменты"]),
                float(row["Сумма для распределения"]),
                row["Специалист/Ресурс.Выполнение"],
                row["Персонал. Дополнительный персонал/ресурсы"],
            ]

            for c, v in enumerate(values, 1):
                cell = ws.cell(r, c, value=v)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if c == 4:   # Дата
                    cell.number_format = "dd.mm.yyyy"
                if c == 5:   # Кол-во
                    cell.number_format = "0"
                if c in (6,7,8):  # суммы
                    cell.number_format = "0.00"
            r += 1

    # Итоги
    total_qty = int(table_df["Количество услуг"].sum()) if not table_df.empty else 0
    total_sum = float(table_df["Сумма по тарифу"].sum()) if not table_df.empty else 0.0
    total_meds = float(table_df["Медикаменты"].sum()) if not table_df.empty else 0.0
    total_dist = float(table_df["Сумма для распределения"].sum()) if not table_df.empty else 0.0

    ws.cell(r, 1, "ИТОГО").font = Font(bold=True)
    ws.cell(r, 5, total_qty).font = Font(bold=True)
    ws.cell(r, 6, round(total_sum, 2)).font = Font(bold=True)
    ws.cell(r, 7, round(total_meds, 2)).font = Font(bold=True)
    ws.cell(r, 8, round(total_dist, 2)).font = Font(bold=True)

    for c in range(1, ncols + 1):
        ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)

    # Границы и формат
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for rr in range(1, r + 1):
        for cc in range(1, ncols + 1):
            ws.cell(rr, cc).border = border

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def build_report(df: pd.DataFrame, output_path: str) -> None:
    meds_map, profile_map = load_db_maps()

    df_prof, df_other = split_by_vid(df)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_med = wb.create_sheet(title=SHEET_MEDCOM)
    write_patient_like_sheet(ws_med, df_prof, meds_map, profile_map)

    ws_other = wb.create_sheet(title=SHEET_OTHER)
    write_patient_like_sheet(ws_other, df_other, meds_map, profile_map)

    wb.save(output_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Входной файл xlsx")
    ap.add_argument("--output", required=True, help="Выходной файл xlsx")
    args = ap.parse_args()

    df = load_input(args.input)
    build_report(df, args.output)
    print(f"Отчёт сохранён: {args.output}")


if __name__ == "__main__":
    main()