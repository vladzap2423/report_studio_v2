# -*- coding: utf-8 -*-
"""
ОТЧЁТ ДЛЯ plat_po_spec
Медикаменты и Профиль берутся ИЗ БАЗЫ data.db
"""
import argparse
import os
import sqlite3

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ====================== ПУТЬ К БАЗЕ ======================
DB_PATH = os.path.normpath(os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "data.db"
))

SHEET_MEDCOM = "Медкомиссии"
SHEET_OTHER = "Остальное"

# Входные колонки
VID_COL = "Вид поступления"
VID_PROF = "профосмотр"

COL_CODE = "Код ОКМУ"
COL_SERVICE = "Услуга"
COL_STATE = "Состояние"
COL_DATE = "Дата"
COL_SUM = "Сумма"
COL_FIO = "ФИО"
COL_SPEC = "Специалист/Ресурс.Выполнение"
COL_EXTRA = "Понятие dopPers"
COL_TARPLAN = "Тар.план"
COL_DOGOVOR = "Договор на оплату"

# Выходные колонки
OUT_HEADERS = [
    "ФИО", "Услуга", "Состояние", "Тар.план", "Договор на оплату",
    "Дата договора", "Количество услуг", "Сумма по тарифу",
    "Медикаменты", "Сумма для распределения",
    "Специалист/Ресурс.Выполнение",
    "Персонал. Дополнительный персонал/ресурсы",
]

COL_WIDTHS = [32, 64, 18, 16, 18, 14, 16, 16, 16, 22, 32, 32]


def get_db_connection():
    if not os.path.exists(DB_PATH):
        raise FileNotFoundError(f"База данных не найдена: {DB_PATH}")
    return sqlite3.connect(DB_PATH)


def load_medicaments_and_profiles():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT "Код_ОК_МУ", "Медикаменты", "Профиль"
        FROM services
        WHERE "Код_ОК_МУ" IS NOT NULL AND "Код_ОК_МУ" != ''
    """)
    rows = cursor.fetchall()
    conn.close()

    meds = {}
    prof = {}
    for code, m, p in rows:
        code = str(code).strip()
        if code:
            meds[code] = float(m) if m is not None else 0.0
            prof[code] = str(p).strip() if p else "Без профиля"
    return meds, prof


def load_input(path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header_row = None
    headers = []
    for r in range(1, 30):
        row_values = [str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        if "Код ОКМУ" in row_values:
            header_row = r
            headers = row_values
            break

    if header_row is None:
        raise RuntimeError("Не найдена строка с заголовком 'Код ОКМУ'")

    data = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {headers[c-1]: ws.cell(r, c).value for c in range(1, len(headers)+1)}
        if all(v is None for v in row.values()):
            continue
        data.append(row)

    df = pd.DataFrame(data)

    df[COL_SUM] = pd.to_numeric(df.get(COL_SUM), errors="coerce").fillna(0)

    for col in [VID_COL, COL_FIO, COL_SERVICE, COL_CODE, COL_STATE, COL_SPEC, COL_EXTRA, COL_TARPLAN, COL_DOGOVOR]:
        if col in df.columns:
            df[col] = df[col].astype("string").str.strip().replace({"nan": None, "None": None, "": None})

    if COL_DATE in df.columns:
        df[COL_DATE] = pd.to_datetime(df[COL_DATE], errors="coerce")

    return df[df[COL_CODE].notna()].copy()


def split_by_vid(df: pd.DataFrame):
    if VID_COL not in df.columns:
        return df.iloc[0:0].copy(), df.copy()
    
    vid_norm = df[VID_COL].str.lower().str.strip()
    df_prof = df[vid_norm == VID_PROF].copy()
    df_other = df[vid_norm != VID_PROF].copy()
    return df_prof, df_other


def aggregate_for_patient_style(df: pd.DataFrame, meds_map: dict, profile_map: dict):
    df = df.copy()

    # Заполняем пропуски
    for col in [COL_FIO, COL_SERVICE, COL_STATE, COL_SPEC, COL_EXTRA, COL_TARPLAN, COL_DOGOVOR]:
        if col in df.columns:
            df[col] = df[col].fillna("")

    df["Профиль"] = df[COL_CODE].astype(str).str.strip().map(profile_map).fillna("Без профиля")

    g = df.groupby(
        ["Профиль", COL_FIO, COL_CODE, COL_SERVICE, COL_STATE, COL_DATE, COL_TARPLAN, COL_DOGOVOR, COL_SPEC, COL_EXTRA],
        as_index=False, dropna=False
    ).agg(
        **{
            "Количество услуг": (COL_CODE, "size"),
            "Сумма по тарифу": (COL_SUM, "sum"),
        }
    )

    per_one = g[COL_CODE].astype(str).str.strip().map(meds_map).fillna(0.0)
    g["Медикаменты"] = (per_one * g["Количество услуг"]).round(2)
    g["Сумма для распределения"] = (g["Сумма по тарифу"] - g["Медикаменты"]).round(2)

    out = pd.DataFrame({
        "Профиль": g["Профиль"],
        "ФИО": g[COL_FIO],
        "Услуга": g[COL_SERVICE],
        "Состояние": g[COL_STATE],
        "Тар.план": g[COL_TARPLAN],
        "Договор на оплату": g[COL_DOGOVOR],
        "Дата договора": g[COL_DATE],
        "Количество услуг": g["Количество услуг"].astype(int),
        "Сумма по тарифу": g["Сумма по тарифу"].round(2),
        "Медикаменты": g["Медикаменты"].round(2),
        "Сумма для распределения": g["Сумма для распределения"].round(2),
        "Специалист/Ресурс.Выполнение": g[COL_SPEC],
        "Персонал. Дополнительный персонал/ресурсы": g[COL_EXTRA],
    })

    out.sort_values(["Профиль", "ФИО", "Тар.план", "Договор на оплату", "Дата договора"], inplace=True)
    return out


def write_patient_like_sheet(ws, df_part: pd.DataFrame, meds_map: dict, profile_map: dict):
    table = aggregate_for_patient_style(df_part, meds_map, profile_map)
    ncols = len(OUT_HEADERS)

    # ЗАЩИТА ОТ NA — КРИТИЧЕСКОЕ ИСПРАВЛЕНИЕ
    table = table.fillna("")
    table = table.replace([None, pd.NA], "")

    # Заголовки
    for c, h in enumerate(OUT_HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    r = 2
    for profile, group in table.groupby("Профиль", sort=False):
        # Заголовок профиля — ИСПРАВЛЕННЫЙ ВЫЗОВ
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        cell = ws.cell(r, 1, profile or "Без профиля")
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")
        r += 1

        for _, row in group.iterrows():
            values = [
                row["ФИО"],
                row["Услуга"],
                row["Состояние"],
                row["Тар.план"],
                row["Договор на оплату"],
                row["Дата договора"],
                int(row["Количество услуг"]) if row["Количество услуг"] else 0,
                float(row["Сумма по тарифу"]) if row["Сумма по тарифу"] else 0.0,
                float(row["Медикаменты"]) if row["Медикаменты"] else 0.0,
                float(row["Сумма для распределения"]) if row["Сумма для распределения"] else 0.0,
                row["Специалист/Ресурс.Выполнение"],
                row["Персонал. Дополнительный персонал/ресурсы"]
            ]
            for c, v in enumerate(values, 1):
                cell = ws.cell(r, c, v)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if c == 6 and v: cell.number_format = "dd.mm.yyyy"
                if c == 7: cell.number_format = "0"
                if c in (8, 9, 10): cell.number_format = "0.00"
            r += 1

    # Итог
    r += 1
    ws.cell(r, 1, "ИТОГО").font = Font(bold=True)
    ws.cell(r, 7, int(table["Количество услуг"].sum() or 0)).font = Font(bold=True)
    ws.cell(r, 8, round(table["Сумма по тарифу"].sum() or 0, 2)).font = Font(bold=True)
    ws.cell(r, 9, round(table["Медикаменты"].sum() or 0, 2)).font = Font(bold=True)
    ws.cell(r, 10, round(table["Сумма для распределения"].sum() or 0, 2)).font = Font(bold=True)

    # Границы и ширины
    thin = Side(style="thin")
    for rr in range(1, r + 1):
        for cc in range(1, ncols + 1):
            ws.cell(rr, cc).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def build_report(df: pd.DataFrame, output_path: str):
    meds_map, profile_map = load_medicaments_and_profiles()

    df_prof, df_other = split_by_vid(df)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_patient_like_sheet(wb.create_sheet(SHEET_MEDCOM), df_prof, meds_map, profile_map)
    write_patient_like_sheet(wb.create_sheet(SHEET_OTHER), df_other, meds_map, profile_map)

    wb.save(output_path)
    print(f"Отчёт успешно создан: {output_path}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    df = load_input(args.input)
    build_report(df, args.output)


if __name__ == "__main__":
    main()